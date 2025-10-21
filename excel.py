import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import os

# =========================================
# 1️⃣ KONFIGURASI
# =========================================
excel_path = "hasil_deteksi_video/hasil_video.csv"
output_excel = "hasil_deteksi_video/list_plat_dengan_gambar.xlsx"

# =========================================
# 2️⃣ BACA CSV (bukan Excel)
# =========================================
df = pd.read_csv(excel_path, sep=";")

# Normalisasi nama kolom (hapus spasi dan jadikan lowercase)
df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")

# Pastikan kolom yang dibutuhkan ada
required_cols = ["crop_path", "hd_path"]
for col in required_cols:
    if col not in df.columns:
        raise KeyError(f"❌ Kolom '{col}' tidak ditemukan di CSV! Kolom yang ada: {df.columns.tolist()}")

# =========================================
# 3️⃣ BUAT WORKBOOK BARU
# =========================================
wb = Workbook()
ws = wb.active
ws.title = "Deteksi Plat"

# Tulis header
for idx, col in enumerate(df.columns, start=1):
    ws.cell(row=1, column=idx, value=col)

# =========================================
# 4️⃣ ISI DATA & TAMBAH GAMBAR
# =========================================
for i, row in df.iterrows():
    excel_row = i + 2  # baris ke-2 ke bawah
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=excel_row, column=j, value=row[col])

    # Tambah gambar crop
    crop_path = str(row["crop_path"]).strip()
    if os.path.exists(crop_path):
        try:
            img = ExcelImage(crop_path)
            img.width, img.height = 120, 60
            ws.add_image(img, f"{chr(65 + df.columns.get_loc('crop_path'))}{excel_row}")
        except Exception as e:
            print(f"⚠️ Gagal menambahkan gambar crop: {e}")

    # Tambah gambar HD
    hd_path = str(row["hd_path"]).strip()
    if os.path.exists(hd_path):
        try:
            img = ExcelImage(hd_path)
            img.width, img.height = 120, 60
            ws.add_image(img, f"{chr(65 + df.columns.get_loc('hd_path'))}{excel_row}")
        except Exception as e:
            print(f"⚠️ Gagal menambahkan gambar HD: {e}")

# =========================================
# 5️⃣ SIMPAN HASIL
# =========================================
wb.save(output_excel)
print(f"✅ File Excel baru dengan gambar tersimpan di:\n{output_excel}")
